import os
import streamlit as st
import pandas as pd
import pg8000
from io import BytesIO
import zipfile
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Customizing the page
st.set_page_config(
    page_title="Generate Theory Exam Pass/Fail Report",
    page_icon="📘",
    layout="wide"
)

# Inline CSS for styling
page_style = """
    <style>
    /* Background color */
    .stApp {
        background-color: #e0e0e0; /* Light grey background */
    }

    /* Text customization */
    .title {
        color: #2b2b2b; /* Dark grey */
        font-size: 42px;
        font-weight: bold;
        text-align: center;
        font-family: 'Arial', sans-serif;
        margin-top: 20px;
    }

    .subtitle {
        color: #595959; /* Medium grey */
        font-size: 24px;
        text-align: center;
        font-family: 'Arial', sans-serif;
        margin-bottom: 40px;
    }

    /* Neumorphic button styling */
    .stButton>button {
        background: #e0e0e0; /* Match the background for a flat look */
        border: none;
        border-radius: 12px;
        box-shadow: 4px 4px 6px #bebebe, -4px -4px 6px #ffffff;
        font-size: 16px;
        padding: 15px 30px;
        font-family: 'Arial', sans-serif;
        color: #2b2b2b; /* Dark grey text */
        cursor: pointer;
    }

    .stButton>button:hover {
        background: #d6d6d6;
        box-shadow: inset 2px 2px 4px #bebebe, inset -2px -2px 4px #ffffff;
        color: #2b2b2b;
    }
    </style>
"""
st.markdown(page_style, unsafe_allow_html=True)

# Title and subtitle
st.markdown('<p class="title">Generate Theory Exam Pass/Fail Report</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Generate an Excel Pass/Fail Report within the specified period</p>', unsafe_allow_html=True)

# "Return to Main Page" Button
if st.button("Return to Main Page"):
    st.experimental_set_query_params(page="main")
    st.experimental_rerun()

# Set a simple password
PASSWORD = os.environ["APP_PASSWORD"]

# Create a password input field in Streamlit
password = st.text_input("Enter Password", type="password")
if password != PASSWORD:
    st.warning("Incorrect password")
    st.stop()
else:
    st.success("Access granted!")

# Function to query database and generate coversheets in a zip file
def generate_coversheets_zip(curriculum, startdate, enddate):
    db_connection = pg8000.connect(
        database=os.environ["SUPABASE_DB_NAME"],
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        host=os.environ["SUPABASE_HOST"],
        port=os.environ["SUPABASE_PORT"]
    )

    db_cursor = db_connection.cursor()

    # Format query to select data within specified date range and curriculum
    db_query = f"""SELECT student_list.name,                  
                    student_list.iatc_id, 
                    student_list.nat_id,
                    student_list.class,
                    student_list.curriculum,
                    exam_results.session,
                    exam_results.exam,
                    exam_results.result,
                    exam_results.date,
                    exam_results.type
                    FROM exam_results 
                    JOIN student_list ON exam_results.nat_id = student_list.nat_id
                    WHERE student_list.curriculum = '{curriculum}' 
                    AND exam_results.date >= '{startdate}' 
                    AND exam_results.date <= '{enddate}'
                    ORDER BY 
                    exam_results.date ASC, 
                    exam_results.session ASC, 
                    student_list.class ASC,
                    student_list.iatc_id ASC
                """
    db_cursor.execute(db_query)
    output_data = db_cursor.fetchall()
    db_cursor.close()
    db_connection.close()

    # Convert output to DataFrame for Excel export
    col_names = ['Name', 'IATC ID', 'National ID', 'Class', 'Curriculum', 'Session', 'Exam', 'Result', 'Date', 'Exam Type']
    df = pd.DataFrame(output_data, columns=col_names)

    # Create an in-memory ZIP file to store individual Excel files
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        # Save the entire DataFrame to one Excel file with formatting
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Pass_Fail_Report")
            worksheet = writer.sheets["Pass_Fail_Report"]

            # Apply header styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set auto-filters on the header row
            worksheet.auto_filter.ref = worksheet.dimensions

            # Autosize columns based on content
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max_length + 2  # Add padding for readability

            # Apply borders and center alignment to all cells
            thin_border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin")
            )
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

        # Save formatted Excel file in the zip
        excel_buffer.seek(0)
        zip_file.writestr(f"{curriculum} Report ({startdate} to {enddate}).xlsx", excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer

# Curriculum selection
curriculum = st.selectbox("Select Faculty:", ["EASA", "GACA", "UAS"])

# Date input from user
startdate = st.date_input("Select Start Date:")
enddate = st.date_input("Select End Date:")

# Button to generate and download coversheets
if st.button("Generate Report"):
    if startdate > enddate:
        st.error("Error: End Date must be after Start Date.")
    else:
        try:
            st.write("Generating Report...")

            # Generate the zip file in memory
            zip_file = generate_coversheets_zip(curriculum, startdate, enddate)

            # Download button for the zip file
            st.download_button(
                label="Download Report",
                data=zip_file,
                file_name="Pass_Fail_Report.zip",
                mime="application/zip"
            )
            st.success("Pass Fail Report zip generated successfully!")
        except Exception as e:
            st.error(f"An error occurred: {e}")
