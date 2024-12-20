
import os
import streamlit as st
import pandas as pd
import pg8000
from io import BytesIO
import zipfile
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Customizing the page
st.set_page_config(
    page_title="Pass-Fail Reporting Portal",
    page_icon="📘",
    layout="wide"
)

# CSS styling
page_style = """
    <style>
    .stApp {
        background-color: #e0e0e0;
    }
    .title {
        color: #2b2b2b;
        font-size: 42px;
        font-weight: bold;
        text-align: center;
        font-family: 'Arial', sans-serif;
    }
    .subtitle {
        color: #595959;
        font-size: 24px;
        text-align: center;
        font-family: 'Arial', sans-serif;
        margin-bottom: 20px;
    }
    </style>
"""
st.markdown(page_style, unsafe_allow_html=True)

# Title
st.markdown('<p class="title">Generate Pass/Fail Reports</p>', unsafe_allow_html=True)

# Set a simple password
PASSWORD = os.environ["APP_PASSWORD1"]
password = st.text_input("Enter Password", type="password")
if password != PASSWORD:
    st.warning("Incorrect password")
    st.stop()

st.success("Access granted!")

# Function to query the database and generate reports
def generate_coversheets_zip(curriculum, startdate, enddate):
    db_connection = pg8000.connect(
        database=os.environ["SUPABASE_DB_NAME"],
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        host=os.environ["SUPABASE_HOST"],
        port=os.environ["SUPABASE_PORT"]
    )
    db_cursor = db_connection.cursor()

    db_query = f"""
        SELECT student_list.name, student_list.iatc_id, student_list.nat_id,
               student_list.class, student_list.curriculum, exam_results.session,
               exam_results.exam, exam_results.result, exam_results.date, exam_results.type
        FROM exam_results
        JOIN student_list ON exam_results.nat_id = student_list.nat_id
        WHERE student_list.curriculum = '{curriculum}'
        AND exam_results.date >= '{startdate}'
        AND exam_results.date <= '{enddate}'
        ORDER BY exam_results.date ASC, exam_results.session ASC, 
                 student_list.class ASC, student_list.iatc_id ASC
    """
    db_cursor.execute(db_query)
    output_data = db_cursor.fetchall()
    db_cursor.close()
    db_connection.close()

    # Convert output to DataFrame for Excel export
    col_names = ['Name', 'IATC ID', 'National ID', 'Class', 'Curriculum', 'Session', 'Exam', 'Result', 'Date', 'Exam Type']
    df = pd.DataFrame(output_data, columns=col_names)

    # Create an in-memory ZIP file to store the report
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Pass_Fail_Report")
            worksheet = writer.sheets["Pass_Fail_Report"]

            # Apply header styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set auto-filters and adjust column width
            worksheet.auto_filter.ref = worksheet.dimensions
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max_length + 2

        # Save Excel file to zip
        excel_buffer.seek(0)
        zip_file.writestr(f"{curriculum} Report ({startdate} to {enddate}).xlsx", excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer

# Curriculum selection and date inputs
curriculum = st.selectbox("Select Faculty:", ["EASA", "GACA", "UAS"])
startdate = st.date_input("Select Start Date:")
enddate = st.date_input("Select End Date:")

# Generate Report Button
if st.button("Generate Report"):
    if startdate > enddate:
        st.error("Error: End Date must be after Start Date.")
    else:
        st.write("Generating report...")
        try:
            zip_file = generate_coversheets_zip(curriculum, startdate, enddate)
            st.download_button(
                label="Download Report",
                data=zip_file,
                file_name="Pass_Fail_Report.zip",
                mime="application/zip"
            )
            st.success("Pass/Fail Report generated successfully!")
        except Exception as e:
            st.error(f"An error occurred: {e}")
