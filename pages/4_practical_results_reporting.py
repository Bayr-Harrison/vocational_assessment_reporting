import streamlit as st
import pandas as pd
import pg8000
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set page configuration
st.set_page_config(
    page_title="Practical Task Results Reporting Portal",
    page_icon="📘",
    layout="wide"
)

# CSS styling
page_style = """
    <style>
    .stApp {
        background-color: #e0e0e0;
    }
    .title {
        color: #2b2b2b;
        font-size: 42px;
        font-weight: bold;
        text-align: center;
        font-family: 'Arial', sans-serif;
    }
    .subtitle {
        color: #595959;
        font-size: 24px;
        text-align: center;
        font-family: 'Arial', sans-serif';
        margin-bottom: 40px;
    }
    </style>
"""
st.markdown(page_style, unsafe_allow_html=True)

# Title
st.markdown('<p class="title">Practical Task Results Reporting Portal</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Query practical task results and download them as an Excel file.</p>', unsafe_allow_html=True)

# Authentication
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    password = st.text_input("Enter Password", type="password")
    if password:
        if password == os.environ["APP_PASSWORD4"]:
            st.session_state["authenticated"] = True
            st.success("Authenticated successfully!")
        else:
            st.error("Incorrect password. Please try again.")

if st.session_state["authenticated"]:
    st.write("Select a date range to query the database and download results as an Excel file.")

    # Date inputs
    start_date = st.date_input("Start Date")
    end_date = st.date_input("End Date")

    # Connect to Supabase database
    def connect_to_supabase():
        return pg8000.connect(
            database=os.environ["SUPABASE_DB_NAME"],
            user=os.environ["SUPABASE_USER"],
            password=os.environ["SUPABASE_PASSWORD"],
            host=os.environ["SUPABASE_HOST"],
            port=os.environ["SUPABASE_PORT"]
        )

    # Query the database
    def query_database(start_date, end_date):
        db_query = f"""
            SELECT 
                prac_results.batch AS "Batch Number",
                prac_results.type AS "Batch Type",
                prac_results.iatc_id AS "IATC ID",
                student_list.name AS "Name",
                student_list.class AS "Class",
                prac_results.task AS "Task",
                prac_results.score AS "Score",
                prac_results.status AS "Attendance Status",
                prac_results.task_date AS "Task Date",
                prac_results.sent_date AS "Sending Date"
            FROM prac_results
            JOIN student_list ON prac_results.iatc_id = student_list.iatc_id
            WHERE prac_results.sent_date >= '{start_date}' 
            AND prac_results.sent_date <= '{end_date}'
            ORDER BY 
                prac_results.task_date ASC, 
                student_list.class ASC, 
                prac_results.iatc_id ASC;
        """
        try:
            connection = connect_to_supabase()
            cursor = connection.cursor()
            cursor.execute(db_query)
            result = cursor.fetchall()
            cursor.close()
            connection.close()
            return result
        except Exception as e:
            st.error(f"Error querying database: {e}")
            return []

    # Create and download Excel file
    def create_excel(data, start_date, end_date):
        col_names = [
            'Batch Number', 'Batch Type', 'IATC ID', 'Name', 'Class',
            'Task', 'Score', 'Attendance Status', 'Task Date', 'Sending Date'
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Practical Task Results"

        # Write headers
        for col_num, header in enumerate(col_names, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = PatternFill(start_color="8DE7D3", end_color="8DE7D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Apply filters and borders
        ws.auto_filter.ref = ws.dimensions
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(col_names)):
            for cell in row:
                cell.border = thin_border

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Button to query and download
    if st.button("Query and Download"):
        if start_date and end_date:
            st.write("Querying database...")
            data = query_database(start_date, end_date)
            if data:
                st.write(f"Query returned {len(data)} rows.")
                excel_file = create_excel(data, start_date, end_date)
                st.download_button(
                    label="Download Excel File",
                    data=excel_file,
                    file_name=f"Practical_Results_{start_date}_to_{end_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("No data found for the specified date range.")
        else:
            st.error("Please select both start and end dates.")
