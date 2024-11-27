import os
import streamlit as st
import pandas as pd
import pg8000
from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Set page configuration
st.set_page_config(
    page_title="Coversheet Generator",
    page_icon="📘",
    layout="wide"
)

# CSS styling
page_style = """
    <style>
    .stApp {
        background-color: #e0e0e0;
    }
    .title {
        color: #2b2b2b;
        font-size: 42px;
        font-weight: bold;
        text-align: center;
        font-family: 'Arial', sans-serif;
    }
    .subtitle {
        color: #595959;
        font-size: 24px;
        text-align: center;
        font-family: 'Arial', sans-serif';
        margin-bottom: 40px;
    }
    </style>
"""
st.markdown(page_style, unsafe_allow_html=True)

# Title and subtitle
st.markdown('<p class="title">Generate Coversheets</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">Create individualized coversheets summarizing each student\'s exam performance.</p>', unsafe_allow_html=True)

# Password authentication
PASSWORD = os.environ["APP_PASSWORD3"]
password = st.text_input("Enter Password", type="password")
if password != PASSWORD:
    st.warning("Incorrect password")
    st.stop()
else:
    st.success("Access granted!")

# Function to establish a database connection
def get_database_connection():
    return pg8000.connect(
        database=os.environ["SUPABASE_DB_NAME"],
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        host=os.environ["SUPABASE_HOST"],
        port=os.environ["SUPABASE_PORT"]
    )

# Function to query the database and fetch data
def fetch_student_data(student_list):
    connection = get_database_connection()
    cursor = connection.cursor()
    
    student_list_string = ', '.join(map(str, student_list))
    query = f"""
        SELECT student_list.name, student_list.iatc_id, student_list.nat_id,
               student_list.class, exam_list.exam_long AS subject,
               exam_results.score, exam_results.result, exam_results.date
        FROM exam_results 
        JOIN student_list ON exam_results.nat_id = student_list.nat_id
        JOIN exam_list ON exam_results.exam = exam_list.exam
        WHERE student_list.iatc_id IN ({student_list_string}) AND exam_results.score_index = 1
        ORDER BY exam_list.srt_exam ASC
    """
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    connection.close()
    col_names = ['Name', 'IATC ID', 'National ID', 'Class', 'Subject', 'Score', 'Result', 'Date']
    return pd.DataFrame(data, columns=col_names)

# Function to create Excel coversheets
def create_coversheet(student_data, student_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Coversheet"

    # Styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="AEE1F8", end_color="AEE1F8", fill_type="solid")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Static labels
    ws["B2"], ws["B3"], ws["B4"], ws["B5"] = "Student Name:", "Student IATC ID:", "Student National ID:", "Student Class:"
    ws["C2"], ws["C3"], ws["C4"], ws["C5"] = student_data.iloc[0, 0], student_data.iloc[0, 1], student_data.iloc[0, 2], student_data.iloc[0, 3]

    # Apply styles to static cells
    for cell in ["B2", "B3", "B4", "B5"]:
        ws[cell].font = header_font
        ws[cell].alignment = cell_alignment
        ws[cell].fill = header_fill
        ws[cell].border = thin_border

    for cell in ["C2", "C3", "C4", "C5"]:
        ws[cell].alignment = cell_alignment
        ws[cell].border = thin_border

    # Headers
    headers = ['Subject', 'Score', 'Result', 'Date']
    for col_num, header in enumerate(headers, 2):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        cell.border = thin_border

    # Data rows
    for row_num, row in enumerate(student_data[['Subject', 'Score', 'Result', 'Date']].values, 8):
        for col_num, value in enumerate(row, 2):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Adjust column widths
    for col in range(2, 6):
        max_length = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    # Save to buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# Generate ZIP file
def generate_coversheets_zip(student_list):
    df = fetch_student_data(student_list)
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for student_id in student_list:
            student_data = df[df['IATC ID'] == student_id]
            excel_buffer = create_coversheet(student_data, student_id)
            zip_file.writestr(f"{student_id}_Coversheet.xlsx", excel_buffer.read())
    zip_buffer.seek(0)
    return zip_buffer

# User input
st.write("Enter student IDs separated by commas:")
student_ids_input = st.text_area("Student IDs (e.g., 151596, 156756, 154960):")
if st.button("Generate Coversheets"):
    try:
        student_list = [int(id.strip()) for id in student_ids_input.split(",")]
        zip_file = generate_coversheets_zip(student_list)
        st.download_button(
            label="Download Coversheets as ZIP",
            data=zip_file,
            file_name="Coversheets.zip",
            mime="application/zip"
        )
        st.success("Coversheets generated successfully!")
    except Exception as e:
        st.error(f"An error occurred: {e}")
